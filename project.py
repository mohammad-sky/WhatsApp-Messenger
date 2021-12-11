from os import name
import time
import glob
from alright import WhatsApp
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import *
import sys
from PyQt5 import QtCore, QtWidgets ,QtGui
from PyQt5.QtWidgets import QMainWindow, QWidget, QLabel, QLineEdit ,QRadioButton ,QPushButton 
# from PyQt5.QtWidgets import QPushButton
from PyQt5.QtCore import QSize    
import openpyxl
from openpyxl.cell import cell 
import os



class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setWindowIcon(QtGui.QIcon("ArmNikan.png"))


        self.setMinimumSize(QSize(1200, 600))    
        self.setWindowTitle("Nikan masenger") 
        self.setStyleSheet("background: Blue Turquoise")  




        self.nameLabel1 = QLabel(self)
        self.nameLabel1.setText('text')
        self.nameLabel1.setFont(QFont('Arial', 15))

        self.line1 = QLineEdit(self)
        self.line1.move(150, 30)
        self.line1.resize(1000, 128)
        self.nameLabel1.move(50, 80)
        self.line1.setStyleSheet("background-color: Violet")


        self.nameLabel2 = QLabel(self)
        self.nameLabel2.setText('address')
        self.nameLabel2.setFont(QFont('Arial', 14))

        self.line2 = QLineEdit(self)
        self.line2.move(150, 220)
        self.line2.resize(1000, 64)
        self.nameLabel2.move(20, 232)
        self.line2.setStyleSheet("background-color: Violet")





        #button of group
        a = QPushButton('جیم', self)
        a.clicked.connect(self.j_group)
        a.resize(150,50)
        a.move(250, 450)     
        a.setStyleSheet("background-color:  lightblue")

        
        b = QPushButton('ب', self)
        b.clicked.connect(self.b_group)
        b.resize(150,50)
        b.move(500, 450)   
        b.setStyleSheet("background-color:  lightblue")

        
        j = QPushButton('الف', self)
        j.clicked.connect(self.a_group)
        j.resize(150,50)
        j.move(750, 450)      
        j.setStyleSheet("background-color:  lightblue")




        path = "name_of_ masenger.xlsx"
        wb_obj = openpyxl.load_workbook(path) 
        self.sheet_obj = wb_obj.active 
        




        
        
        radio_button1 = QRadioButton(self)
        # setting geometry of radio button
        radio_button1.setGeometry(150, 333, 100, 40)
        # setting text to radio button
        radio_button1.setText("Image")
    
  
        radio_button2 = QRadioButton(self)
        # setting geometry of radio button
        radio_button2.setGeometry(250, 333, 100, 40)
        # setting text to radio button
        radio_button2.setText("Video")  
        
        
        
        radio_button3 = QRadioButton(self)
        # setting geometry of radio button
        radio_button3.setGeometry(350, 333, 100, 40)
        # setting text to radio button
        radio_button3.setText("file")
        
        
        radio_button4 = QRadioButton(self)
        # setting geometry of radio button
        radio_button4.setGeometry(430, 333, 100, 40)
        # setting text to radio button
        radio_button4.setText("mp3") 
  
  
        
        
        
        cs1 = QRadioButton("Send a file to everyone",self)
        cs1.setGeometry(850, 333, 200, 40)

        cs2 = QRadioButton("Based on numbering",self)
        cs2.setGeometry(650, 333, 200, 40)

        # cs3 = QRadioButton("Expert",self)
        # cs3.move(130, 60)

        cs_group = QButtonGroup(self)
        cs_group.addButton(cs1)
        cs_group.addButton(cs2)
        # cs_group.addButton(cs3)

                
                
                
        
        
        
        pdf = radio_button3
        self.pdf = pdf


        Image = radio_button1
        self.Image = Image

        Video = radio_button2
        self.Video = Video
      
      
        mp3 = radio_button4
        self.mp3 = mp3
        
        Based_on_numbering = cs2
        self.Based_on_numbering = Based_on_numbering
          
          
        to_everyone = cs1
        self.to_everyone = to_everyone
        
        
        self.show()

        
        
    def j_group(self):
        messenger = WhatsApp()
        if self.Based_on_numbering.isChecked() ==True:
            if self.pdf.isChecked() ==True:
                # list_1=[glob.glob(f"{self.line2.text()}/*.jpg")]
                list_1=[]
                list_1.append(glob.glob(f"{self.line2.text()}/*.docx"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.pptx"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.xlsx"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.pdf"))
                list_2=list_1[0]
                
                

                
                names =[]
                for i in range(2):
                    cell_obj = self.sheet_obj.cell(row = (i+1), column = 2)
                    names.append(cell_obj.value)

                time.sleep(20)
                for i in names:
                    for r in list_1:
                        for x in r:   
                            string=os.path.basename(x)
                            string1=os.path.splitext(string)[0]
                            

                            if f"{string1}"==f"{(names.index(i))+1}":

                                messenger.find_by_username(i)
                                messenger.send_file(x)
                                messenger.send_message(self.line1.text()) 
                                time.sleep(5)   
                time.sleep(30)
                # messenger.logout()
                
                
                
                
            elif self.Image.isChecked() ==True:
                # list_1=[glob.glob(f"{self.line2.text()}/*.jpg")]
                list_1=[]
                list_1.append(glob.glob(f"{self.line2.text()}/*.jpg"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.png"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.svg"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.webp"))

                list_2=list_1[0]

                names =[]
                for i in range(2):
                    cell_obj = self.sheet_obj.cell(row = (i+1), column = 2)
                    names.append(cell_obj.value)

                time.sleep(20)
                for i in names:
                    for r in list_1:
                        for x in r:   
                            string=os.path.basename(x)
                            string1=os.path.splitext(string)[0]
                            

                            if f"{string1}"==f"{(names.index(i))+1}":

                                messenger.find_by_username(i)
                                messenger.send_picture(x)
                                messenger.send_message(self.line1.text()) 
                                time.sleep(7)   
                time.sleep(45)
                # messenger.logout()          
                
                
                
                
            elif self.Video.isChecked() ==True:
                list_1=[]
                list_1.append(glob.glob(f"{self.line2.text()}/*.mp4"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.avi"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.mov"))
                list_1.append(glob.glob(f"{self.line2.text()}/*.wmv"))    
                        
                list_2=list_1[0]
                
                
                names =[]
                for i in range(2):
                    cell_obj = self.sheet_obj.cell(row = (i+1), column = 2)
                    names.append(cell_obj.value)

                time.sleep(20)
                for i in names:
                    for r in list_1:
                        for x in r:   
                            string=os.path.basename(x)
                            string1=os.path.splitext(string)[0]
                            

                            if f"{string1}"==f"{(names.index(i))+1}":

                                messenger.find_by_username(i)
                                messenger.send_video(x)
                                messenger.send_message(self.line1.text()) 
                                time.sleep(10)   
                time.sleep(45)
                # messenger.logout()          
                
                
                
            elif self.mp3.isChecked() ==True:
                list_1=[glob.glob(f"{self.line2.text()}/*.mp3")]
                list_2=list_1[0]
                
                
                
                
                names =[]
                for i in range(2):
                    cell_obj = self.sheet_obj.cell(row = (i+1), column = 2)
                    names.append(cell_obj.value)

                time.sleep(20)
                for i in names:
                    for r in list_1:
                        for x in r:   
                            string=os.path.basename(x)
                            string1=os.path.splitext(string)[0]
                            

                            if f"{string1}"==f"{(names.index(i))+1}":

                                messenger.find_by_username(i)
                                messenger.send_file(x)
                                messenger.send_message(self.line1.text()) 
                                time.sleep(10)   
                time.sleep(45)
                # messenger.logout()      
                
                
                
                
                
                
        elif self.to_everyone.isChecked() ==True:
            pass         
                
            







    def b_group(self):
        messenger = WhatsApp()

        if self.pdf.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.pdf")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_file(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(10)   
            time.sleep(30)
            # messenger.logout()
        elif self.Image.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.jpg")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_picture(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(10)   
            time.sleep(45)
            # messenger.logout()            
        elif self.Video.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.mp4")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_video(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(15)   
            time.sleep(60)
            # messenger.logout() 
        elif self.mp3.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.mp3")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_file(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(15)   
            time.sleep(60)
            # messenger.logout()             
            
                    
    def a_group(self):
        messenger = WhatsApp()

        if self.pdf.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.pdf")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_file(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(5)   
            time.sleep(30)
            # messenger.logout()
        if self.Image.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.jpg")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_picture(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(10)   
            time.sleep(45)
            # messenger.logout()            
        elif self.Video.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.mp4")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_video(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(15)   
            time.sleep(60)
            # messenger.logout() 
        elif self.mp3.isChecked() ==True:
            list_1=[glob.glob(f"{self.line2.text()}/*.mp3")]
            list_2=list_1[0]
            names =["Negar","Davod"]
            time.sleep(20)
            for i in names:
                for x in list_2:
                    if x.count(i)==1:
                        messenger.find_by_username(i)
                        messenger.send_file(x)
                        messenger.send_message(self.line1.text()) 
                        time.sleep(15)   
            time.sleep(60)
            # messenger.logout()             
            
                    
            
  



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    mainWin = MainWindow()
    mainWin.show()
    sys.exit( app.exec_() )
